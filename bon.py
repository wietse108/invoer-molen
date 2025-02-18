import tkinter as tk
from tkinter import ttk
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
import os

# Excel bestand instellen
EXCEL_FILE = "bonnen.xlsx"

def maak_excel():
    """Maak de Excel-sheet aan als deze niet bestaat."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Graan Bonnen"
        ws.append(["ID", "Graansoort", "Gewicht", "Datum", "THT"])
        wb.save(EXCEL_FILE)
    print("Excel-bestand is ingesteld.")

# Functie: Nieuwe bon opslaan in de Excel-sheet
def sla_bon_op(graansoort, gewicht, datum, tht):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    new_id = ws.max_row  # Nieuwe ID is de volgende rij
    ws.append([new_id, graansoort, gewicht, datum, tht])
    wb.save(EXCEL_FILE)

# Functie: Cijfers toevoegen aan input
def voeg_cijfer_toe(cijfer):
    huidige_waarde = bon_label_var.get()
    if huidige_waarde == "25":
        bon_label_var.set(cijfer)
    else:
        bon_label_var.set(huidige_waarde + cijfer)

# Functie: Bon informatie tonen en opslaan
def maak_bon():
    try:
        graansoort = graansoort_var.get()
        if not graansoort:
            raise ValueError("Kies een graansoort.")
        gewicht = float(bon_label_var.get())
        vandaag = datetime.now()
        datum_str = vandaag.strftime("%Y-%m-%d")
        tht = vandaag + timedelta(days=6 * 30)  # THT = 6 maanden later
        tht_str = tht.strftime("%Y-%m-%d")

        # Opslaan in de Excel-sheet
        sla_bon_op(graansoort, gewicht, datum_str, tht_str)

        # Boninformatie in de terminal tonen
        print(f"Bon gemaakt: Graansoort={graansoort}, Gewicht={gewicht} kg, Datum={datum_str}, THT={tht_str}")

        # Reset
        bon_label_var.set("25")
    except ValueError as e:
        print(f"Fout: {str(e)}")

# Functie: Afsluiten
def afsluiten():
    root.attributes('-fullscreen', False)
    root.destroy()

# GUI Setup
root = tk.Tk()
root.title("Graan Bon Systeem")
root.attributes('-fullscreen', True)  # Fullscreen-modus
root.bind("<Escape>", lambda e: afsluiten())  # Escape sluit fullscreen

# Graansoort selecteren
graansoort_var = tk.StringVar()
graansoorten = [
    "Tarwe", "BIO Tarwe",
    "Spelt", "BIO Spelt",
    "Rogge", "BIO Rogge"
]

tk.Label(root, text="Kies een graansoort:", font=("Arial", 16)).pack(pady=10)
graansoort_menu = ttk.Combobox(root, textvariable=graansoort_var, values=graansoorten, state="readonly", font=("Arial", 16))
graansoort_menu.pack(pady=5)

# Hoofdframe in het midden van het scherm
main_frame = tk.Frame(root)
main_frame.place(relx=0.5, rely=0.5, anchor="center")  # Centreer het frame

# Layout frames binnen het hoofdframe
left_frame = tk.Frame(main_frame)
left_frame.grid(row=0, column=0, padx=20, pady=20)

right_frame = tk.Frame(main_frame)
right_frame.grid(row=0, column=1, padx=20, pady=20)

# Gewicht label en Bon-knop
bon_label_var = tk.StringVar(value="25")  # Gewicht standaard op 25 kg
bon_label = tk.Label(right_frame, textvariable=bon_label_var, font=("Arial", 36), bg="#f0f0f0", width=10, height=2, relief="solid")
bon_label.pack(pady=10)

tk.Button(right_frame, text="Bon", font=("Arial", 24), bg="#4caf50", fg="white", width=10, height=2, command=maak_bon).pack(pady=10)

# Nummerieke toetsen
keypad = [
    ('7', 0, 0), ('8', 0, 1), ('9', 0, 2),
    ('4', 1, 0), ('5', 1, 1), ('6', 1, 2),
    ('1', 2, 0), ('2', 2, 1), ('3', 2, 2),
    ('0', 3, 1)
]
for key, row, col in keypad:
    tk.Button(left_frame, text=key, font=("Arial", 24), width=4, height=2, bg="#2196f3", fg="white",
              command=lambda k=key: voeg_cijfer_toe(k)).grid(row=row, column=col, padx=5, pady=5)

# Start de applicatie
maak_excel()
root.mainloop()
