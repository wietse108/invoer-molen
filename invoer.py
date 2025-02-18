import tkinter as tk
from tkinter import messagebox, ttk
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

# Excel bestand instellen

EXCEL_FILE = "graan_data.xlsx"
BACKGROUND_COLOR = "#f0f0f0"  # Light grey background
BUTTON_COLOR = "#4a76a8"       # Main button color
BUTTON_TEXT_COLOR = "white"
BUTTON_ACTIVE_COLOR = "#355c7d" # Darker color when clicked
LABEL_COLOR = "#333"           # Dark text color
FONT_NORMAL = ("Arial", 14)    # Regular font
FONT_LARGE = ("Arial", 18, "bold")  # Large heading font
FONT_NUMERIC = ("Arial", 24) # Numeric keypad font

# Functie: Maak de Excel-sheet aan als deze niet bestaat
def maak_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Graan Inputs"
        ws.append(["ID", "Gewicht", "Graansoort", "Datum", "THT"])
        wb.save(EXCEL_FILE)
    print("Excel-bestand is ingesteld.")

# Functie: Gegevens opslaan in de Excel-sheet
def sla_op_in_excel(gewicht, graansoort, datum, tht):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    new_id = ws.max_row  # Nieuwe ID is de volgende rij
    ws.append([new_id, gewicht, graansoort, datum, tht])
    wb.save(EXCEL_FILE)

# Functie: Gegevens ophalen van vandaag
def haal_inputs_vandaag_op():
    """Haal alle inputs van vandaag op uit de Excel-sheet."""
    vandaag = datetime.now().strftime('%Y-%m-%d')
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] == vandaag:
            data.append(row[1:4])  # Gewicht, Graansoort, Datum
    return data

# Functie: Gegevens berekenen en opslaan
def bereken_en_opslaan():
    try:
        gewicht = float(entry_gewicht.get())
        graansoort = graansoort_var.get()
        if not graansoort:
            raise ValueError("Kies een graansoort.")
        vandaag = datetime.now()
        datum_str = vandaag.strftime("%Y-%m-%d")
        tht = vandaag + timedelta(days=6 * 30)  # THT = 6 maanden later
        tht_str = tht.strftime("%Y-%m-%d")

        # Opslaan in de Excel-sheet
        sla_op_in_excel(gewicht, graansoort, datum_str, tht_str)

        messagebox.showinfo("Succes", f"{graansoort} van {gewicht:.2f} kg is opgeslagen!\nTHT: {tht_str}")
        wis_invoer()
        update_tabel()  # Update de tabel met de nieuwe invoer
    except ValueError as e:
        messagebox.showerror("Fout", str(e))

# Functie: Wis de invoer velden
def wis_invoer():
    entry_gewicht.delete(0, tk.END)
    graansoort_var.set(None)

# Functie: Update de tabel met gegevens van vandaag
def update_tabel():
    for row in tree.get_children():
        tree.delete(row)
    data = haal_inputs_vandaag_op()
    for item in data:
        tree.insert("", "end", values=item)

# Functie: Voeg een cijfer toe aan het gewicht invoerveld
def voeg_cijfer_toe(cijfer):
    current_value = entry_gewicht.get()
    entry_gewicht.delete(0, tk.END)
    entry_gewicht.insert(0, current_value + cijfer)

# Functie: Verwijder de laatste invoer
def verwijder_laatste_invoer():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    max_row = ws.max_row
    if max_row > 1:
        ws.delete_rows(max_row)
        wb.save(EXCEL_FILE)
        update_tabel()
        messagebox.showinfo("Succes", "Laatste invoer is ongedaan gemaakt.")
    else:
        messagebox.showinfo("Info", "Er is geen invoer om ongedaan te maken.")

# GUI Setup
root = tk.Tk()
root.title("Graanmolen Bonnetjes")
root.attributes('-fullscreen', True)  # Fullscreen modus
root.configure(bg=BACKGROUND_COLOR)  # Set global background

def afsluiten():
    root.attributes('-fullscreen', False)
    root.destroy()

root.bind("<Escape>", lambda e: afsluiten())  # Escape sluit fullscreen

# Hoofdscherm met twee frames (links voor invoer, rechts voor tabel)
main_frame = tk.Frame(root, bg=BACKGROUND_COLOR)
main_frame.pack(fill="both", expand=True)

# Linker frame (invoer scherm)
left_frame = tk.Frame(main_frame, bg=BACKGROUND_COLOR)
left_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)

# Rechter frame (tabel)
right_frame = tk.Frame(main_frame, bg=BACKGROUND_COLOR)
right_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)

# Titel
tk.Label(left_frame, text="Graanmolen Bonnetjes", font=("Arial", 30, "bold"), bg=BACKGROUND_COLOR, fg=LABEL_COLOR).pack(pady=20)

# Gewicht invoeren
tk.Label(left_frame, text="Voer het gewicht van graan in (kg):", font=FONT_NORMAL, bg=BACKGROUND_COLOR, fg=LABEL_COLOR).pack(pady=10)
entry_gewicht = tk.Entry(left_frame, justify="right", font=FONT_LARGE, width=10, bd=2, relief="solid")
entry_gewicht.pack(pady=10)

# Graansoort selecteren
tk.Label(left_frame, text="Kies het graansoort:", font=FONT_NORMAL, bg=BACKGROUND_COLOR, fg=LABEL_COLOR).pack(pady=20)

graansoort_frame = tk.Frame(left_frame, bg=BACKGROUND_COLOR)
graansoort_frame.pack(pady=10)

graansoort_var = tk.StringVar(value=None)
graansoorten = [
    "Tarwe", "BIO Tarwe", 
    "Spelt", "BIO Spelt", 
    "Rogge", "BIO Rogge"
]
graansoort_knoppen = {}

# Functie: Update de graansoort knoppen
def update_graansoort_knoppen():
    for knop, soort in graansoort_knoppen.items():
        if graansoort_var.get() == soort:
            knop.config(relief=tk.SUNKEN)
        else:
            knop.config(relief=tk.RAISED)

# Maak de graansoortknoppen
for index, soort in enumerate(graansoorten):
    knop = tk.Button(
        graansoort_frame, text=soort, width=10, height=2, font=FONT_NORMAL, bg=BUTTON_COLOR, fg=BUTTON_TEXT_COLOR,
        activebackground=BUTTON_ACTIVE_COLOR, activeforeground="white", relief=tk.GROOVE,
        command=lambda s=soort: (graansoort_var.set(s), update_graansoort_knoppen())
    )
    knop.grid(row=index // 3, column=index % 3, padx=10, pady=10)
    graansoort_knoppen[knop] = soort

# Numeriek toetsenbord frame
keypad_frame = tk.Frame(left_frame, bg=BACKGROUND_COLOR)
keypad_frame.pack(pady=20)

# Numerieke toetsen
cijfers = [
    ('7', 0, 0), ('8', 0, 1), ('9', 0, 2),
    ('4', 1, 0), ('5', 1, 1), ('6', 1, 2),
    ('1', 2, 0), ('2', 2, 1), ('3', 2, 2),
    ('0', 3, 1)
]

for cijfer, rij, kolom in cijfers:
    tk.Button(
        keypad_frame,
        text=cijfer,
        width=5,
        height=2,
        font=FONT_NUMERIC,
        bg="#4caf50",  # Green color
        fg="white",
        activebackground="#45a049",
        activeforeground="white",
        relief="raised", # Raised appearance
        command=lambda c=cijfer: voeg_cijfer_toe(c)
    ).grid(row=rij, column=kolom, padx=5, pady=5)

# Wissen en OK-knoppen
tk.Button(
    keypad_frame,
    text="C",
    width=5,
    height=2,
    font=FONT_NUMERIC,
    bg="#f44336",  # Red color
    fg="white",
    activebackground="#e53935",
    activeforeground="white",
    relief="raised",
    command=wis_invoer
).grid(row=3, column=0, padx=5, pady=5)

tk.Button(
    keypad_frame,
    text="OK",
    width=5,
    height=2,
    font=FONT_NUMERIC,
    bg="#2196f3",  # Blue color
    fg="white",
    activebackground="#1e88e5",
    activeforeground="white",
    relief="raised",
    command=bereken_en_opslaan
).grid(row=3, column=2, padx=5, pady=5)

# Tabel voor de dagelijkse inputs (rechts)
tk.Label(right_frame, text="Dagelijkse Inputs", font=FONT_LARGE, bg=BACKGROUND_COLOR, fg=LABEL_COLOR).pack(pady=10)

# Treeview voor de tabel
style = ttk.Style()
style.configure("Treeview", font=FONT_NORMAL, background=BACKGROUND_COLOR, foreground=LABEL_COLOR)
style.configure("Treeview.Heading", font=FONT_NORMAL, background=BACKGROUND_COLOR, foreground=LABEL_COLOR)
tree = ttk.Treeview(right_frame, columns=("gewicht", "graansoort", "datum"), show="headings", height=20, style="Treeview")
tree.heading("gewicht", text="Gewicht (kg)")
tree.heading("graansoort", text="Graansoort")
tree.heading("datum", text="Datum")
tree.column("gewicht", anchor="center", width=100)
tree.column("graansoort", anchor="center", width=150)
tree.column("datum", anchor="center", width=150)
tree.pack(pady=10, fill="both", expand=True)

# Undo knop onder de tabel
undo_button = tk.Button(
    right_frame,
    text="Undo Laatste Invoer",
    font=FONT_NORMAL,
    bg="#f44336",  # Rood
    fg="white",
    activebackground="#e53935",
    activeforeground="white",
    relief="raised",
    command=verwijder_laatste_invoer
)
undo_button.pack(pady=10)

# Start de GUI
maak_excel()
update_tabel()  # Vul de tabel met gegevens van vandaag
root.mainloop()